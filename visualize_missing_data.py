import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter

INPUT_FILE = "author_metadata_with_occ.xlsx"
OUTPUT_FILE = "author_metadata_final.xlsx"

def main():
    print(f"Loading {INPUT_FILE}...")
    try:
        wb = openpyxl.load_workbook(INPUT_FILE)
        ws = wb.active
    except FileNotFoundError:
        print(f"Error: {INPUT_FILE} not found.")
        return

    # Find column indices
    headers = [cell.value for cell in ws[1]]
    print(f"Headers found in {INPUT_FILE}: {headers}")
    
    col_map = {name: i+1 for i, name in enumerate(headers)}
    
    # Identify key columns (adjust names based on actual file)
    # Expected: 'Work QID'/'QID', 'Name'/'Author Name', 'Birth'/'Birth Year', 'Death'/'Death Year', 'Floruit'
    # The refined script output 'Birth', 'Death', 'Floruit'. The first script output 'Birth Year' etc.
    # Let's check what we have. Refined/augmented script was using refined output which had 'Birth', 'Death', 'Floruit'.
    
    birth_col = col_map.get('Birth') or col_map.get('Birth Year')
    death_col = col_map.get('Death') or col_map.get('Death Year')
    floruit_col = col_map.get('Floruit')
    
    if not (birth_col and death_col and floruit_col):
        print("Error: Could not find Date columns (Birth, Death, Floruit).")
        print("Columns found:", headers)
        return

    # Insert WD_status column. Let's put it at index 2 (Column B), shifting everything else.
    # Actually, inserting columns in openpyxl can be slow or tricky with references, 
    # but for simple data it works.
    # User suggestion: "補助列を1本追加する... e.g. WD_status"
    
    wd_status_col_idx = 2 # Column B
    ws.insert_cols(wd_status_col_idx)
    ws.cell(row=1, column=wd_status_col_idx).value = "WD_status"
    
    # Adjust valid col indices after insertion if they were >= 2
    # But wait, insert_cols shifts existing columns.
    # So if Birth was 3, it becomes 4.
    # Re-calculate or just iterate headers again? 
    # Better to just insert and then re-map.
    
    # Re-map headers
    headers = [cell.value for cell in ws[1]] # This might not update immediately if we only changed cell value?
    # Actually insert_cols shifts.
    
    # Let's simple iterate rows to populate data
    # Row 1 is header.
    
    # Re-find columns by name (old names are shifted)
    # The new column is at `wd_status_col_idx` (B).
    # The old columns are now shifted by 1 if they were >= 2.
    
    # Let's just assume we insert at B.
    # QID (A) -> A.
    # Name (B) -> C.
    # Birth (C) -> D. etc.
    
    # Instead of assumption, let's look for headers dynamically again.
    # But headers inside the sheet are shifted? No, header row content shifts automatically?
    # No, insert_cols inserts empty column. We set the value of B1.
    # The previous B1 content moves to C1.
    
    # Populate WD_status
    print("Populating WD_status...")
    
    # We need to know where Birth, Death, Floruit are NOW.
    # Recalculate map
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    col_map = {name: i for i, name in enumerate(headers, 1)}
    
    birth_col = col_map.get('Birth') or col_map.get('Birth Year')
    death_col = col_map.get('Death') or col_map.get('Death Year')
    floruit_col = col_map.get('Floruit')
    
    wd_stat_letter = get_column_letter(wd_status_col_idx) # B
    
    for row in range(2, ws.max_row + 1):
        b_val = ws.cell(row=row, column=birth_col).value
        d_val = ws.cell(row=row, column=death_col).value
        f_val = ws.cell(row=row, column=floruit_col).value
        
        # Check if empty (None or empty string)
        def is_empty(v):
            return v is None or str(v).strip() == ""
            
        all_missing = is_empty(b_val) and is_empty(d_val) and is_empty(f_val)
        
        status_val = "Wikidata欠損" if all_missing else "Wikidataあり"
        ws.cell(row=row, column=wd_status_col_idx).value = status_val

    # Conditional Formatting
    print("Applying conditional formatting...")
    
    # Fills
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light Yellow
    text_red = Font(color="9C0006")
    text_green = Font(color="006100")
    bg_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
    
    # Rule 1: Empty Date Cells -> Red
    # Range: e.g. D2:D1000
    for col_idx in [birth_col, death_col, floruit_col]:
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}2:{col_letter}{ws.max_row}"
        
        # ISBLANK doesn't always catch empty strings from pandas.
        # Use LEN(TRIM(Cell))=0
        # formula = f'LEN(TRIM({col_letter}2))=0' 
        # Actually standard ISBLANK is safest for strictly empty, but let's stick to user request "セルの値 → 空白"
        # openpyxl 'blanks' rule.
        
        rule = CellIsRule(operator='equal', formula=['""'], stopIfTrue=False, fill=red_fill)
        # Or simple expression rule for blanks?
        # Let's use formula: =LEN(TRIM(D2))=0
        
        # In openpyxl, formula rules need the top-left cell relative ref.
        rule = FormulaRule(formula=[f'LEN(TRIM({col_letter}2))=0'], stopIfTrue=False, fill=red_fill)
        ws.conditional_formatting.add(rng, rule)

    # Rule 2: Entire Row Highlight if WD_status="Wikidata欠損"
    # Range: A2:Z1000 (entire data)
    full_range = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    
    # Formula needs to lock column B ($B2)
    # =$B2="Wikidata欠損"
    # Note: Using get_column_letter matching wd_status_col_idx
    status_ref = f"${wd_stat_letter}2"
    rule_row = FormulaRule(formula=[f'{status_ref}="Wikidata欠損"'], stopIfTrue=False, fill=yellow_fill)
    ws.conditional_formatting.add(full_range, rule_row)
    
    # Rule 3: WD_status text color
    # Range: B2:B1000
    status_rng = f"{wd_stat_letter}2:{wd_stat_letter}{ws.max_row}"
    
    rule_stat_missing = CellIsRule(operator='equal', formula=['"Wikidata欠損"'], stopIfTrue=False, font=text_red, fill=red_fill)
    ws.conditional_formatting.add(status_rng, rule_stat_missing)
    
    rule_stat_present = CellIsRule(operator='equal', formula=['"Wikidataあり"'], stopIfTrue=False, font=text_green, fill=bg_green)
    ws.conditional_formatting.add(status_rng, rule_stat_present)
    
    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done.")

if __name__ == "__main__":
    main()
